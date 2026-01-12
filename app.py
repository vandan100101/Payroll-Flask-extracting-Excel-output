# ============================================================================
# FLASK PAYROLL PROCESSING APPLICATION
# Modern web-based replacement for VBA Excel payroll system
# ============================================================================

from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
from werkzeug.utils import secure_filename
import traceback
import numpy as np

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls', 'xlsb', 'txt'}

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ============================================================================
# CONFIGURATION & CONSTANTS
# ============================================================================

COMPANY_NAME = "Leader Electronics (Philippine Branch) Incorporated"
MIN_WAGE = "Minimum Wage      P600.00/day"
YEAR = 2025

# Month configurations
MONTH_CONFIG = {
    'January': {'code': '01', 'prev': 'December', 'days': 31},
    'February': {'code': '02', 'prev': 'January', 'days': 28},
    'March': {'code': '03', 'prev': 'February', 'days': 31},
    'April': {'code': '04', 'prev': 'March', 'days': 30},
    'May': {'code': '05', 'prev': 'April', 'days': 31},
    'June': {'code': '06', 'prev': 'May', 'days': 30},
    'July': {'code': '07', 'prev': 'June', 'days': 31},
    'August': {'code': '08', 'prev': 'July', 'days': 31},
    'September': {'code': '09', 'prev': 'August', 'days': 30},
    'October': {'code': '10', 'prev': 'September', 'days': 31},
    'November': {'code': '11', 'prev': 'October', 'days': 30},
    'December': {'code': '12', 'prev': 'November', 'days': 31}
}

# Cost Center Mapping
CCR_CODE_MAPPING = {
    'A': 'IND2001',
    'B': 'IND2005',
    'C': 'IND2101',
    'D': 'IND2102',
    'E': 'IND0202',
    'F': 'IND0202-1',
    'G': 'IND0203',
    'H': 'IND0203-1',
    'I': 'IND0204',
    'J': 'IND0205',
    'K': 'IND0503',
    'L': 'IND0506',
    'M': 'IND0702',
    'N': 'D2001',
    'O': 'D2005',
    'P': 'IND1002'
}

# Reverse mapping
CCR_NAME_TO_CODE = {v: k for k, v in CCR_CODE_MAPPING.items()}

# Department color codes (hex)
DEPT_COLORS = {
    'IND_PROD': 'CDFFCF',        # Light Green
    'IND_QA': 'D1FFFF',          # Light Cyan
    'IND_QA_ALT': 'FFFFD5',      # Light Yellow
    'IND_WAREHOUSE': 'FFEBFF',   # Light Pink
    'IND_702': 'CCFFFF',         # Light Blue
    'DIRECT_PROD': 'FFF1DD',     # Light Orange
    'IND_1002': 'CCCCFF',        # Light Purple
    'GRAND_TOTAL': 'FABF8F'      # Peach
}

# Department mappings
DEPT_TOTALS = {
    1: 'TOTAL IND2001',
    2: 'TOTAL IND2005',
    3: 'TOTAL IND2101',
    4: 'TOTAL IND2102',
    5: 'TOTAL IND202',
    6: 'TOTAL IND202-1',
    7: 'TOTAL IND203',
    8: 'TOTAL IND203-1',
    9: 'TOTAL IND204',
    10: 'TOTAL IND205',
    11: 'TOTAL IND503',
    12: 'TOTAL IND506',
    13: 'TOTAL IND702',
    14: 'TOTAL D2001',
    15: 'TOTAL D2005',
    16: 'TOTAL IND1002'
}

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def safe_int(value, default=0):
    """Safely convert to integer"""
    try:
        return int(float(value)) if pd.notna(value) else default
    except (ValueError, TypeError):
        return default

def safe_float(value, default=0.0):
    """Safely convert to float"""
    try:
        return float(value) if pd.notna(value) else default
    except (ValueError, TypeError):
        return default

# ============================================================================
# PAYROLL PROCESSING CLASS
# ============================================================================

class PayrollProcessor:
    def __init__(self, dataframe, dbase_df, month, cutoff):
        self.df = dataframe
        self.dbase = dbase_df
        self.month = month
        self.cutoff = cutoff
        self.month_info = MONTH_CONFIG[month]
        
    def add_lookups(self):
        """Add CCR code and account number lookups"""
        try:
            # Add CCR CODE lookup (column 6 from dbase)
            self.df['CCR_CODE'] = self.df.iloc[:, 0].map(
                self.dbase.set_index(self.dbase.columns[0])[self.dbase.columns[5]]
            ).fillna('Not in dbase')
            
            # Add Account Number lookup (column 4 from dbase)
            self.df['ACCT_NO'] = self.df.iloc[:, 0].map(
                self.dbase.set_index(self.dbase.columns[0])[self.dbase.columns[3]]
            ).fillna('Not in dbase')
            
            # Reorder columns to put lookups at front
            cols = ['CCR_CODE', self.df.columns[0], 'ACCT_NO'] + \
                   [col for col in self.df.columns if col not in ['CCR_CODE', 'ACCT_NO', self.df.columns[0]]]
            self.df = self.df[cols]
            
        except Exception as e:
            print(f"Error in add_lookups: {e}")
            raise
    
    def sort_data(self):
        """Sort by CCR code, then by employee ID"""
        try:
            self.df = self.df.sort_values(
                by=[self.df.columns[0], self.df.columns[1], self.df.columns[3]],
                na_position='last'
            )
            self.df.reset_index(drop=True, inplace=True)
        except Exception as e:
            print(f"Error in sort_data: {e}")
            raise
    
    def add_13th_month(self):
        """Add 13th month pay calculation"""
        try:
            # Assuming column 8 is basic salary
            if len(self.df.columns) > 7:
                self.df['13TH_MONTH'] = self.df.iloc[:, 7].apply(
                    lambda x: safe_float(x) / 12.0
                )
        except Exception as e:
            print(f"Error in add_13th_month: {e}")
            raise
    
    def insert_subtotals(self):
        """Insert subtotal rows by department"""
        try:
            print(f"  Initial dataframe shape: {self.df.shape}")
            print(f"  Grouping by column: {self.df.columns[0]}")
            
            # Group by CCR code
            grouped = self.df.groupby(self.df.columns[0], dropna=False)
            print(f"  Found {len(grouped)} groups")
            
            result_dfs = []
            dept_counter = 1
            
            # Track accumulated totals for group summaries
            ind_prod_groups = []
            ind_qa_groups = []
            ind_warehouse_groups = []
            direct_prod_groups = []
            
            # Track all employee data rows for grand total
            all_employee_rows = []
            
            for name, group in grouped:
                if pd.isna(name):
                    continue
                
                print(f"  Processing group '{name}': {len(group)} employees")
                
                # Store employee data rows (not subtotals)
                all_employee_rows.append(group)
                    
                # Add the group data
                result_dfs.append(group)
                
                # Create subtotal row
                subtotal_row = pd.Series([''] * len(self.df.columns), index=self.df.columns)
                subtotal_label = DEPT_TOTALS.get(dept_counter, f'TOTAL {name}')
                subtotal_row[self.df.columns[0]] = name  # Keep CCR code
                subtotal_row[self.df.columns[2]] = subtotal_label
                subtotal_row[self.df.columns[1]] = len(group)  # Count
                
                # Sum numeric columns
                numeric_cols_summed = 0
                for col in self.df.columns[7:]:
                    if self.df[col].dtype in ['int64', 'float64']:
                        subtotal_row[col] = group[col].sum()
                        numeric_cols_summed += 1
                
                print(f"    Subtotal: {subtotal_label}, summed {numeric_cols_summed} numeric columns")
                
                result_dfs.append(pd.DataFrame([subtotal_row]))
                
                # Track for group totals
                if dept_counter in [1, 2]:
                    ind_prod_groups.append(subtotal_row)
                elif dept_counter in [3, 4, 5, 6, 7, 8, 9, 10]:
                    ind_qa_groups.append(subtotal_row)
                elif dept_counter in [11, 12]:
                    ind_warehouse_groups.append(subtotal_row)
                elif dept_counter in [14, 15]:
                    direct_prod_groups.append(subtotal_row)
                
                # Add special group totals with spacing
                if dept_counter == 2:  # After IND2005 (B)
                    group_total = self._create_group_total(ind_prod_groups, 'IND PROD TOTAL')
                    result_dfs.append(pd.DataFrame([group_total]))
                    blank_row = pd.Series([''] * len(self.df.columns), index=self.df.columns)
                    result_dfs.append(pd.DataFrame([blank_row]))
                    print(f"    Added IND PROD TOTAL and spacing")
                    
                elif dept_counter == 10:  # After IND205 (J)
                    group_total = self._create_group_total(ind_qa_groups, 'IND QA TOTAL')
                    result_dfs.append(pd.DataFrame([group_total]))
                    blank_row = pd.Series([''] * len(self.df.columns), index=self.df.columns)
                    result_dfs.append(pd.DataFrame([blank_row]))
                    print(f"    Added IND QA TOTAL and spacing")
                    
                elif dept_counter == 12:  # After IND506 (L)
                    group_total = self._create_group_total(ind_warehouse_groups, 'IND WAREHOUSE TOTAL')
                    result_dfs.append(pd.DataFrame([group_total]))
                    blank_row = pd.Series([''] * len(self.df.columns), index=self.df.columns)
                    result_dfs.append(pd.DataFrame([blank_row]))
                    print(f"    Added IND WAREHOUSE TOTAL and spacing")
                    
                elif dept_counter == 15:  # After D2005 (O)
                    group_total = self._create_group_total(direct_prod_groups, 'DIRECT PROD TOTAL')
                    result_dfs.append(pd.DataFrame([group_total]))
                    blank_row = pd.Series([''] * len(self.df.columns), index=self.df.columns)
                    result_dfs.append(pd.DataFrame([blank_row]))
                    print(f"    Added DIRECT PROD TOTAL and spacing")
                
                dept_counter += 1
            
            # Combine all dataframes
            print(f"\n  Combining {len(result_dfs)} dataframes...")
            self.df = pd.concat(result_dfs, ignore_index=True)
            print(f"  After combining: {self.df.shape}")
            
            # Add GRAND TOTAL DAILY row
            print(f"\n  Creating GRAND TOTAL DAILY...")
            grand_total_row = pd.Series([''] * len(self.df.columns), index=self.df.columns)
            grand_total_row[self.df.columns[2]] = 'GRAND TOTAL DAILY'
            
            # Combine all employee data
            all_employees_df = pd.concat(all_employee_rows, ignore_index=True)
            print(f"  Total employee rows for grand total: {len(all_employees_df)}")
            
            # Sum employee counts and numeric columns from employee rows only
            grand_total_row[self.df.columns[1]] = len(all_employees_df)
            
            # Find which columns have numeric data and sum them
            numeric_cols_in_grand_total = 0
            for col_idx, col in enumerate(self.df.columns):
                if col_idx >= 7:  # Start from column 7 onwards
                    if all_employees_df[col].dtype in ['int64', 'float64']:
                        col_sum = all_employees_df[col].sum()
                        grand_total_row[col] = col_sum
                        numeric_cols_in_grand_total += 1
            
            print(f"  Summed {numeric_cols_in_grand_total} numeric columns in grand total")
            
            self.df = pd.concat([self.df, pd.DataFrame([grand_total_row])], ignore_index=True)
            print(f"  Final dataframe shape: {self.df.shape}")
            
            # Log grand total for verification
            print(f"\n=== Grand Total Verification ===")
            print(f"Total Employees: {len(all_employees_df)}")
            print(f"Number of columns: {len(self.df.columns)}")
            
            # Find Net Pay column
            try:
                net_pay_col_idx = None
                max_sum = 0
                
                print(f"\nSearching for Net Pay column in last 5 columns...")
                for col_idx in range(max(0, len(self.df.columns) - 5), len(self.df.columns)):
                    col = self.df.columns[col_idx]
                    if all_employees_df[col].dtype in ['int64', 'float64']:
                        col_sum = all_employees_df[col].sum()
                        print(f"  Column {col_idx} ({col}): sum = ₱{col_sum:,.2f}")
                        if col_sum > max_sum:
                            max_sum = col_sum
                            net_pay_col_idx = col_idx
                
                if net_pay_col_idx is not None:
                    net_pay_total = safe_float(max_sum)
                    print(f"\n✓ Net Pay found in column {net_pay_col_idx}: ₱{net_pay_total:,.2f}")
                else:
                    print("\n✗ Could not find Net Pay column")
                    
            except Exception as e:
                print(f"\n✗ Error finding Net Pay: {e}")
                print(traceback.format_exc())
            
        except Exception as e:
            print(f"\n✗ ERROR in insert_subtotals:")
            print(f"  Error type: {type(e).__name__}")
            print(f"  Error message: {str(e)}")
            print(traceback.format_exc())
            raise
    
    def _create_group_total(self, group_rows, label):
        """Helper to create group total rows"""
        if not group_rows:
            return pd.Series([''] * len(self.df.columns), index=self.df.columns)
        
        total_row = pd.Series([''] * len(self.df.columns), index=self.df.columns)
        total_row[self.df.columns[2]] = label
        
        # Sum all numeric columns from subtotal rows
        for col in self.df.columns[1:]:
            if col in group_rows[0].index:
                col_sum = sum(safe_float(row[col]) for row in group_rows)
                if col_sum != 0:
                    total_row[col] = col_sum
        
        return total_row
    
    def process(self):
        """Run complete processing pipeline"""
        try:
            print("\n[STEP 1] Adding lookups...")
            self.add_lookups()
            print("✓ Lookups added")
            
            print("\n[STEP 2] Sorting data...")
            self.sort_data()
            print("✓ Data sorted")
            
            print("\n[STEP 3] Adding 13th month pay...")
            self.add_13th_month()
            print("✓ 13th month pay calculated")
            
            print("\n[STEP 4] Inserting subtotals...")
            self.insert_subtotals()
            print("✓ Subtotals inserted")
            
            print(f"\n✓ Processing complete - Final shape: {self.df.shape}")
            return self.df
            
        except Exception as e:
            print(f"\n✗ ERROR in process step:")
            print(f"  Error type: {type(e).__name__}")
            print(f"  Error message: {str(e)}")
            print(traceback.format_exc())
            raise

# ============================================================================
# BDO CONVERTER
# ============================================================================

class BDOConverter:
    def __init__(self, paste_df, dbase_df):
        self.paste_df = paste_df
        self.dbase = dbase_df
        
    def convert(self):
        """Convert payroll data to BDO format"""
        try:
            print(f"\n=== BDO Converter Debug Info ===")
            print(f"Payroll DataFrame shape: {self.paste_df.shape}")
            print(f"Database DataFrame shape: {self.dbase.shape}")
            
            # Show database structure
            print(f"\nDatabase columns: {list(self.dbase.columns)}")
            print(f"Database sample row: {self.dbase.iloc[0].tolist()}")
            
            # Show payroll structure
            print(f"\nPayroll columns (first 10): {list(self.paste_df.columns[:10])}")
            
            bank_data = []
            cash_data = []
            
            # Find Net Pay column
            net_pay_col = None
            
            # Based on typical payroll structure, Net Pay is often in column 33 (AH)
            # But let's search for it more intelligently
            for try_col in [33, 34, 35, 32, 31, 40, 41, 42]:
                if try_col < len(self.paste_df.columns):
                    col_data = self.paste_df.iloc[:, try_col]
                    numeric_data = pd.to_numeric(col_data, errors='coerce')
                    non_zero_values = numeric_data[numeric_data > 0]
                    
                    if len(non_zero_values) > 0:
                        avg_val = non_zero_values.mean()
                        if 1000 < avg_val < 200000:  # Typical net pay range
                            net_pay_col = try_col
                            col_letter = openpyxl.utils.get_column_letter(try_col + 1)
                            print(f"\nFound Net Pay at column {try_col} ({col_letter}) - avg: ₱{avg_val:,.2f}")
                            print(f"Column name: {self.paste_df.columns[try_col] if try_col < len(self.paste_df.columns) else 'Unknown'}")
                            break
            
            if net_pay_col is None:
                # Search all columns
                print(f"\nSearching all columns for Net Pay...")
                for col_idx in range(len(self.paste_df.columns) - 1, max(0, len(self.paste_df.columns) - 30), -1):
                    col_data = self.paste_df.iloc[:, col_idx]
                    numeric_data = pd.to_numeric(col_data, errors='coerce')
                    non_zero_values = numeric_data[numeric_data > 0]
                    
                    if len(non_zero_values) > 10:
                        avg_val = non_zero_values.mean()
                        if 1000 < avg_val < 200000:
                            net_pay_col = col_idx
                            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                            print(f"Found Net Pay at column {col_idx} ({col_letter}) - avg: ₱{avg_val:,.2f}")
                            print(f"Column name: {self.paste_df.columns[col_idx]}")
                            break
            
            if net_pay_col is None:
                # Try to find by column name
                for col_idx, col_name in enumerate(self.paste_df.columns):
                    if isinstance(col_name, str):
                        if 'net' in col_name.lower() or 'pay' in col_name.lower():
                            col_data = self.paste_df.iloc[:, col_idx]
                            numeric_data = pd.to_numeric(col_data, errors='coerce')
                            non_zero_values = numeric_data[numeric_data > 0]
                            if len(non_zero_values) > 10:
                                net_pay_col = col_idx
                                print(f"Found Net Pay by name at column {col_idx}: {col_name}")
                                break
            
            if net_pay_col is None:
                raise ValueError(f"Could not find Net Pay column. Please check your payroll file.")
            
            print(f"\nUsing Net Pay column index: {net_pay_col}")
            
            # Create lookups from database
            account_lookup = {}
            name_lookup = {}
            
            print(f"\nProcessing database records...")
            processed_count = 0
            
            for idx, row in self.dbase.iterrows():
                # DATABASE STRUCTURE:
                # Column 0: Employee ID
                # Column 1: Full Name (already formatted as "Last, First M.")
                # Column 2: Cost Center
                # Column 3: Account Number
                
                emp_id = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                
                # Clean employee ID
                if emp_id and emp_id.replace('.', '').replace('-', '').isdigit():
                    emp_id_clean = emp_id.split('.')[0]
                    
                    # Account number (Column 3)
                    if len(row) > 3 and pd.notna(row.iloc[3]):
                        acct_val = row.iloc[3]
                        
                        # Handle different account number formats
                        if isinstance(acct_val, (int, np.integer)):
                            account_no = str(int(acct_val))
                        elif isinstance(acct_val, float):
                            # Remove decimal if it's .0
                            if acct_val.is_integer():
                                account_no = str(int(acct_val))
                            else:
                                account_no = str(acct_val).split('.')[0]
                        elif isinstance(acct_val, str):
                            # Clean string - keep only digits
                            account_no = ''.join(filter(str.isdigit, acct_val))
                        else:
                            account_no = str(acct_val).split('.')[0]
                        
                        # Validate account number (should be 10-12 digits)
                        if account_no and len(account_no) >= 10:
                            account_lookup[emp_id_clean] = account_no
                        else:
                            if processed_count < 5:  # Show first few warnings
                                print(f"  Warning: Invalid account length for {emp_id_clean}: {account_no}")
                    
                    # Employee name (Column 1 - already formatted!)
                    if len(row) > 1 and pd.notna(row.iloc[1]):
                        employee_name = str(row.iloc[1]).strip()
                        
                        # Clean up name
                        if employee_name and employee_name != 'nan':
                            name_lookup[emp_id_clean] = employee_name
                        else:
                            name_lookup[emp_id_clean] = f"Employee {emp_id_clean}"
                    
                    processed_count += 1
                    if processed_count <= 3:
                        print(f"  DB Record: ID={emp_id_clean}, Name={name_lookup.get(emp_id_clean, 'N/A')}, Account={account_lookup.get(emp_id_clean, 'N/A')}")
            
            print(f"\nCreated lookups:")
            print(f"  - Accounts: {len(account_lookup)} records")
            print(f"  - Names: {len(name_lookup)} records")
            
            # Process payroll rows
            skipped_rows = 0
            bank_count = 0
            cash_count = 0
            processed_emp_ids = set()  # Track processed IDs to prevent duplicates
            skipped_details = {'no_emp_id': 0, 'zero_pay': 0, 'keyword': 0, 'duplicate': 0}
            
            print(f"\nProcessing payroll rows...")
            
            for idx, row in self.paste_df.iterrows():
                try:
                    # Find employee ID from payroll - PAYROLL STRUCTURE:
                    # Column 0: CCR Code (like 'IND2001')
                    # Column 1: Employee ID (numeric)
                    # Column 2: Account Number
                    # Column 3: Last Name
                    # Column 4: First Name
                    
                    emp_id = None
                    
                    # Check column 1 first (most likely employee ID)
                    if len(row) > 1:
                        val = str(row.iloc[1]).strip()
                        if val and val.replace('.', '').replace('-', '').isdigit():
                            emp_id = val.split('.')[0]
                    
                    # If not found, check other columns
                    if not emp_id:
                        for col_idx in [0, 2, 3]:
                            if col_idx < len(row):
                                val = str(row.iloc[col_idx]).strip()
                                if val and val.replace('.', '').replace('-', '').isdigit() and len(val) >= 4:
                                    emp_id = val.split('.')[0]
                                    break
                    
                    if not emp_id:
                        skipped_rows += 1
                        skipped_details['no_emp_id'] += 1
                        continue
                    
                    # Skip if already processed (duplicate)
                    if emp_id in processed_emp_ids:
                        skipped_rows += 1
                        skipped_details['duplicate'] += 1
                        continue
                    
                    # Skip header/total rows
                    row_text = ' '.join(str(x).upper() for x in row.iloc[:5] if pd.notna(x))
                    if any(keyword in row_text for keyword in ['TOTAL', 'GRAND', 'CCR', 'CODE', 'NAME', 'ACCOUNT']):
                        skipped_rows += 1
                        skipped_details['keyword'] += 1
                        continue
                    
                    # Get net pay
                    if net_pay_col < len(row):
                        net_pay = safe_float(row.iloc[net_pay_col])
                    else:
                        net_pay = 0
                    
                    if net_pay <= 0:
                        skipped_rows += 1
                        skipped_details['zero_pay'] += 1
                        continue
                    
                    # Look up account and name
                    account_no = account_lookup.get(emp_id)
                    employee_name = name_lookup.get(emp_id)
                    
                    # If no name in database lookup, build from payroll
                    if not employee_name:
                        name_parts = []
                        # Try to get name from payroll (columns 3, 4, 5)
                        for name_col in [3, 4, 5]:
                            if name_col < len(row):
                                name_part = str(row.iloc[name_col]).strip()
                                if name_part and name_part != 'nan':
                                    name_parts.append(name_part)
                        
                        if len(name_parts) >= 2:
                            # Format as "Last, First Middle"
                            employee_name = f"{name_parts[0]}, {name_parts[1]}"
                            if len(name_parts) > 2 and name_parts[2]:
                                employee_name += f" {name_parts[2][0]}."  # Middle initial
                        elif name_parts:
                            employee_name = name_parts[0]
                        else:
                            employee_name = f"Employee {emp_id}"
                    
                    # Separate into bank or cash
                    if account_no:
                        # HAS BANK ACCOUNT
                        # Clean and format account number
                        account_clean = ''.join(filter(str.isdigit, str(account_no)))
                        
                        # Pad to 10 digits if needed
                        if len(account_clean) < 10:
                            account_clean = account_clean.zfill(10)
                        
                        # Add "00" prefix for BDO
                        account_with_prefix = f"00{account_clean}"
                        
                        bank_data.append({
                            'Account No.': account_with_prefix,
                            'Net Pay': net_pay,
                            'Name': employee_name
                        })
                        
                        bank_count += 1
                        processed_emp_ids.add(emp_id)
                        
                        if bank_count <= 3:
                            print(f"✓ BANK: {emp_id} -> {account_with_prefix}, {employee_name}, ₱{net_pay:,.2f}")
                    else:
                        # NO BANK ACCOUNT - CASH PAYROLL
                        cash_data.append({
                            'Emp ID': emp_id,
                            'Net Pay': net_pay,
                            'Name': employee_name
                        })
                        
                        cash_count += 1
                        processed_emp_ids.add(emp_id)
                        
                        if cash_count <= 3:
                            print(f"💵 CASH: {emp_id} -> {employee_name}, ₱{net_pay:,.2f}")
                    
                except Exception as row_error:
                    print(f"  Error on row {idx}: {row_error}")
                    skipped_rows += 1
                    continue
            
            print(f"\n=== Conversion Summary ===")
            print(f"Total payroll rows processed: {len(self.paste_df)}")
            print(f"Bank payroll employees: {bank_count}")
            print(f"Cash payroll employees: {cash_count}")
            print(f"Skipped rows: {skipped_rows}")
            print(f"  - No employee ID: {skipped_details['no_emp_id']}")
            print(f"  - Zero/negative pay: {skipped_details['zero_pay']}")
            print(f"  - Keyword/header rows: {skipped_details['keyword']}")
            print(f"  - Duplicate entries: {skipped_details['duplicate']}")
            
            if not bank_data and not cash_data:
                raise ValueError(
                    f"No valid employee records found.\n"
                    f"Database has {len(account_lookup)} accounts and {len(name_lookup)} names.\n"
                    f"Check that employee IDs in payroll file match those in database."
                )
            
            # Create dataframes
            bank_df = pd.DataFrame(bank_data) if bank_data else pd.DataFrame(columns=['Account No.', 'Net Pay', 'Name'])
            cash_df = pd.DataFrame(cash_data) if cash_data else pd.DataFrame(columns=['Emp ID', 'Net Pay', 'Name'])
            
            # Calculate totals
            bank_total = bank_df['Net Pay'].sum() if len(bank_df) > 0 else 0
            cash_total = cash_df['Net Pay'].sum() if len(cash_df) > 0 else 0
            total_payroll = bank_total + cash_total
            
            print(f"\n💰 BDO Bank Total: ₱{bank_total:,.2f}")
            print(f"💵 Cash Total: ₱{cash_total:,.2f}")
            print(f"📊 Total Payroll: ₱{total_payroll:,.2f}")
            
            # Sort and clean dataframes
            if len(bank_df) > 0:
                bank_df = bank_df.sort_values('Name').reset_index(drop=True)
            
            if len(cash_df) > 0:
                cash_df = cash_df.sort_values('Name').reset_index(drop=True)
            
            return {
                'bank': bank_df,
                'cash': cash_df,
                'bank_total': bank_total,
                'cash_total': cash_total,
                'total': total_payroll,
                'bank_count': bank_count,
                'cash_count': cash_count
            }
            
        except Exception as e:
            print(f"Error in BDO conversion: {e}")
            traceback.print_exc()
            raise

# ============================================================================
# EXCEL WRITER WITH ENHANCED FORMATTING
# ============================================================================

class FormattedExcelWriter:
    def __init__(self, filename, df, month, cutoff, dbase_df=None):
        self.filename = filename
        self.df = df
        self.month = month
        self.cutoff = cutoff
        self.dbase_df = dbase_df
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = f"{MONTH_CONFIG[month]['code']}{'15' if cutoff == 'first' else '30'}"
        
    def write_headers(self):
        """Write company headers with enhanced formatting"""
        month_info = MONTH_CONFIG[self.month]
        
        if self.cutoff == 'first':
            period_text = f"{month_info['prev']} 26 - {self.month} 10"
            cutoff_text = f"{self.month} 15"
        else:
            period_text = f"{self.month} 11 - 25"
            cutoff_text = f"{self.month} {month_info['days']}"
        
        # Row 1: Company name (merged across columns A-F)
        self.ws.merge_cells('A1:F1')
        self.ws['A1'] = COMPANY_NAME
        self.ws['A1'].font = Font(name='Arial', size=12, bold=True)
        self.ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        self.ws.row_dimensions[1].height = 20
        
        # Row 2: Cut-off period
        self.ws.merge_cells('A2:F2')
        self.ws['A2'] = f"Cut-Off Period:  {period_text}, {YEAR}"
        self.ws['A2'].font = Font(name='Arial', size=10, bold=True)
        self.ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Row 3: Payroll period (red text)
        self.ws.merge_cells('A3:F3')
        self.ws['A3'] = f"Payroll Period:  {cutoff_text}, {YEAR}"
        self.ws['A3'].font = Font(name='Arial', size=10, bold=True, color='FF0000')
        self.ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Row 2 right side: Minimum wage
        self.ws.merge_cells('G2:I2')
        self.ws['G2'] = MIN_WAGE
        self.ws['G2'].font = Font(name='Arial', size=9)
        self.ws['G2'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Row 4: Empty for spacing
        self.ws.row_dimensions[4].height = 10
        
        # Row 5: Empty for spacing
        self.ws.row_dimensions[5].height = 5
        
        # Row 6: Empty for spacing
        self.ws.row_dimensions[6].height = 5
        
        # Add merged header rows for grouping (Row 7)
        self.add_grouped_headers()
        
    def add_grouped_headers(self):
        """Add merged header cells for column grouping"""
        # Row 7 height
        self.ws.row_dimensions[7].height = 20
        
        # Define merged ranges and their labels
        merged_headers = [
            ('O7:Q7', 'DEDUCTION'),
            ('S7:U7', 'EMPLOYEE CONTRIBUTION'),
            ('Y7:AA7', 'NON-TAXABLE EARNINGS'),
            ('AE7:AG7', 'DEDUCTION FOR PAYROLL'),
            ('AJ7:AM7', 'EMPLOYER CONTRIBUTION'),
            ('AP7:AU7', 'Validation'),
            ('AW7:BB7', 'Difference')
        ]
        
        for cell_range, label in merged_headers:
            self.ws.merge_cells(cell_range)
            cell = self.ws[cell_range.split(':')[0]]
            cell.value = label
            cell.font = Font(name='Arial', size=9, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            
            # Add border to merged header cells
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Apply border to all cells in merged range
            start_col = openpyxl.utils.column_index_from_string(cell_range.split(':')[0][0])
            end_col = openpyxl.utils.column_index_from_string(cell_range.split(':')[1][0])
            for col in range(start_col, end_col + 1):
                self.ws.cell(row=7, column=col).border = thin_border
        
    def write_data(self):
        """Write dataframe to Excel with enhanced formatting"""
        # Column headers at row 8
        for col_idx, col_name in enumerate(self.df.columns, start=1):
            cell = self.ws.cell(row=8, column=col_idx, value=col_name)
            cell.font = Font(name='Arial', size=9, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            
            # Add border
            cell.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
        
        # Set row 8 height
        self.ws.row_dimensions[8].height = 50
        
        # Write data starting at row 9
        for row_idx, row_data in enumerate(self.df.itertuples(index=False), start=9):
            row_height = 18  # Default row height
            is_total_row = False
            is_blank_row = False
            
            # Check if this is a blank row (spacing)
            if len(row_data) > 2 and pd.isna(row_data[2]) and pd.isna(row_data[0]):
                is_blank_row = True
                row_height = 10  # Smaller height for spacing rows
            
            # Check if this is a total row
            if len(row_data) > 2 and not is_blank_row:
                row_name = str(row_data[2]) if pd.notna(row_data[2]) else ''
                if any(keyword in row_name for keyword in ['TOTAL', 'GRAND']):
                    is_total_row = True
                    row_height = 22
            
            self.ws.row_dimensions[row_idx].height = row_height
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Replace CCR code letter with full name in column A (except for total rows)
                if col_idx == 1 and not is_total_row and not is_blank_row and value:
                    ccr_code = str(value).strip()
                    if ccr_code in CCR_CODE_MAPPING:
                        cell.value = CCR_CODE_MAPPING[ccr_code]
                
                # Font formatting
                if is_total_row:
                    cell.font = Font(name='Arial', size=10, bold=True)
                elif is_blank_row:
                    cell.font = Font(name='Arial', size=10)
                else:
                    cell.font = Font(name='Arial', size=10)
                
                # Alignment
                cell.alignment = Alignment(vertical='center')
                
                # Center align specific columns (Employee ID column and some others)
                if col_idx in [2, 7]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == 1:  # CCR CODE column
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == 3:  # Account number or name columns
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Number formatting for currency columns (columns H onwards)
                if col_idx >= 8 and isinstance(value, (int, float)) and not is_blank_row:
                    if value == 0:
                        cell.number_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
                    else:
                        cell.number_format = '_-* #,##0.00_-;[Red]-* #,##0.00_-;_-* "-"??_-;_-@_-'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # Add borders (skip for blank rows)
                if not is_blank_row:
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
    
    def apply_department_colors(self):
        """Apply department-specific color coding"""
        for row_idx, row in enumerate(self.df.itertuples(index=False), start=9):
            # Get the department/total name (column 3, index 2)
            row_name = str(row[2]) if len(row) > 2 else ''
            
            color = None
            is_bold = False
            merge_cols = False
            row_height = 18
            
            # Determine color based on row type
            if 'IND2001' in row_name or 'IND2005' in row_name:
                color = DEPT_COLORS['IND_PROD']
                is_bold = True
                merge_cols = True
                row_height = 22
            elif 'IND PROD TOTAL' in row_name:
                color = DEPT_COLORS['IND_PROD']
                is_bold = True
                merge_cols = True
                row_height = 24
            elif 'IND2101' in row_name or 'IND2102' in row_name:
                color = DEPT_COLORS['IND_QA']
                is_bold = True
                merge_cols = True
                row_height = 22
            elif any(x in row_name for x in ['IND202', 'IND203', 'IND204', 'IND205']):
                color = DEPT_COLORS['IND_QA_ALT']
                is_bold = True
                merge_cols = True
                row_height = 22
            elif 'IND QA TOTAL' in row_name:
                color = DEPT_COLORS['IND_QA_ALT']
                is_bold = True
                merge_cols = True
                row_height = 24
            elif 'IND503' in row_name or 'IND506' in row_name:
                color = DEPT_COLORS['IND_WAREHOUSE']
                is_bold = True
                merge_cols = True
                row_height = 22
            elif 'IND WAREHOUSE TOTAL' in row_name:
                color = DEPT_COLORS['IND_WAREHOUSE']
                is_bold = True
                merge_cols = True
                row_height = 24
            elif 'IND702' in row_name:
                color = DEPT_COLORS['IND_702']
                is_bold = True
                merge_cols = True
                row_height = 22
            elif 'D2001' in row_name or 'D2005' in row_name:
                color = DEPT_COLORS['DIRECT_PROD']
                is_bold = True
                merge_cols = True
                row_height = 22
            elif 'DIRECT PROD TOTAL' in row_name:
                color = DEPT_COLORS['DIRECT_PROD']
                is_bold = True
                merge_cols = True
                row_height = 24
            elif 'IND1002' in row_name:
                color = DEPT_COLORS['IND_1002']
                is_bold = True
                merge_cols = True
                row_height = 22
            elif 'GRAND TOTAL' in row_name:
                color = DEPT_COLORS['GRAND_TOTAL']
                is_bold = True
                merge_cols = True
                row_height = 26
            
            # Set row height
            if row_height > 18:
                self.ws.row_dimensions[row_idx].height = row_height
            
            # Apply formatting
            if color:
                for col_idx in range(1, min(41, len(self.df.columns) + 1)):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    
                    if is_bold:
                        cell.font = Font(name='Arial', size=10, bold=True)
                    
                    # Keep number formatting
                    if col_idx >= 8 and isinstance(cell.value, (int, float)):
                        if cell.value == 0:
                            cell.number_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
                        else:
                            cell.number_format = '_-* #,##0.00_-;[Red]-* #,##0.00_-;_-* "-"??_-;_-@_-'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # Merge cells C:E for total rows
                if merge_cols and row_idx <= self.ws.max_row:
                    try:
                        # Unmerge first if already merged
                        self.ws.unmerge_cells(f'C{row_idx}:E{row_idx}')
                    except:
                        pass
                    
                    try:
                        self.ws.merge_cells(f'C{row_idx}:E{row_idx}')
                        merged_cell = self.ws[f'C{row_idx}']
                        merged_cell.alignment = Alignment(horizontal='left', vertical='center')
                    except:
                        pass
    
    def apply_borders(self):
        """Apply professional borders to all data cells"""
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Apply to data range (from row 7 to last row)
        for row in self.ws.iter_rows(min_row=7, max_row=self.ws.max_row, 
                                      min_col=1, max_col=min(40, len(self.df.columns))):
            for cell in row:
                if cell.value is not None or cell.row >= 8:
                    cell.border = thin_border
    
    def set_column_widths(self):
        """Set optimal column widths"""
        column_widths = {
            'A': 8,    # CCR Code
            'B': 7,    # Employee ID
            'C': 12,   # Account No
            'D': 25,   # Employee Name
            'E': 15,   # Department/Position
            'F': 0,    # Hidden column
            'G': 8,    # Center No
            'H': 12,   # Basic Pay
            'I': 10,   # OT columns
            'J': 10,
            'K': 10,
            'L': 10,
            'M': 10,
            'N': 10,
            'O': 10,
            'P': 10,
            'Q': 10,
            'R': 11,   # Total Deduct
            'S': 10,   # SSS
            'T': 10,   # PhilHealth
            'U': 10,   # Pag-IBIG
        }
        
        # Set specific widths
        for col_letter, width in column_widths.items():
            if width == 0:
                self.ws.column_dimensions[col_letter].hidden = True
            else:
                self.ws.column_dimensions[col_letter].width = width
        
        # Set default width for remaining columns (V onwards)
        for col_idx in range(22, 50):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            self.ws.column_dimensions[col_letter].width = 10
    
    def add_signatures(self):
        """Add approval signatures at the bottom"""
        last_row = self.ws.max_row
        
        # Add blank row for spacing after data
        self.ws.row_dimensions[last_row + 1].height = 15
        
        # Add additional info section (Average Days, Med Fee, etc.)
        info_row = last_row + 3
        
        self.ws.cell(row=info_row, column=6).value = "Average Days:"
        self.ws.cell(row=info_row, column=6).font = Font(name='Arial', size=9, bold=True)
        self.ws.cell(row=info_row, column=6).alignment = Alignment(horizontal='left', vertical='center')
        
        # You can add formula here if needed
        # self.ws.cell(row=info_row, column=8).value = formula for average days
        
        # Medical fees section
        self.ws.cell(row=info_row, column=16).value = "Medical Fee:"
        self.ws.cell(row=info_row, column=16).font = Font(name='Arial', size=9, bold=True)
        
        # Benefits section on the right
        benefits_start = info_row
        self.ws.cell(row=benefits_start, column=34).value = "HMI:"
        self.ws.cell(row=benefits_start, column=34).font = Font(name='Arial', size=9, bold=True)
        
        self.ws.cell(row=benefits_start + 1, column=34).value = "Optical:"
        self.ws.cell(row=benefits_start + 1, column=34).font = Font(name='Arial', size=9, bold=True)
        
        self.ws.cell(row=benefits_start + 2, column=34).value = "Dental:"
        self.ws.cell(row=benefits_start + 2, column=34).font = Font(name='Arial', size=9, bold=True)
        
        # Add spacing before signatures
        self.ws.row_dimensions[last_row + 7].height = 20
        self.ws.row_dimensions[last_row + 8].height = 20
        
        # Approval section - well-spaced
        approval_start = last_row + 12
        
        # APPROVED BY section (columns F-H)
        self.ws.merge_cells(f'F{approval_start}:H{approval_start}')
        approved_cell = self.ws[f'F{approval_start}']
        approved_cell.value = "APPROVED BY:"
        approved_cell.font = Font(name='Arial', size=10, bold=True)
        approved_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add blank rows for signature
        self.ws.row_dimensions[approval_start + 1].height = 10
        self.ws.row_dimensions[approval_start + 2].height = 30
        
        # Signature line
        self.ws.merge_cells(f'F{approval_start + 3}:H{approval_start + 3}')
        sig_line = self.ws[f'F{approval_start + 3}']
        sig_line.value = "________________________________"
        sig_line.font = Font(name='Arial', size=10)
        sig_line.alignment = Alignment(horizontal='center', vertical='center')
        
        # Name
        self.ws.merge_cells(f'F{approval_start + 4}:H{approval_start + 4}')
        name_cell = self.ws[f'F{approval_start + 4}']
        name_cell.value = "YEN-PAN HSUEH"
        name_cell.font = Font(name='Arial', size=11, bold=True)
        name_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Position
        self.ws.merge_cells(f'F{approval_start + 5}:H{approval_start + 5}')
        position_cell = self.ws[f'F{approval_start + 5}']
        position_cell.value = "Director"
        position_cell.font = Font(name='Arial', size=10)
        position_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # PREPARED BY section (columns K-M)
        self.ws.merge_cells(f'K{approval_start}:M{approval_start}')
        prepared_cell = self.ws[f'K{approval_start}']
        prepared_cell.value = "PREPARED BY:"
        prepared_cell.font = Font(name='Arial', size=10, bold=True)
        prepared_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Signature line
        self.ws.merge_cells(f'K{approval_start + 3}:M{approval_start + 3}')
        prep_sig_line = self.ws[f'K{approval_start + 3}']
        prep_sig_line.value = "________________________________"
        prep_sig_line.font = Font(name='Arial', size=10)
        prep_sig_line.alignment = Alignment(horizontal='center', vertical='center')
        
        # Name
        self.ws.merge_cells(f'K{approval_start + 4}:M{approval_start + 4}')
        prep_name = self.ws[f'K{approval_start + 4}']
        prep_name.value = "RACQUEL CABRAL"
        prep_name.font = Font(name='Arial', size=11, bold=True)
        prep_name.alignment = Alignment(horizontal='center', vertical='center')
        
        # Position
        self.ws.merge_cells(f'K{approval_start + 5}:M{approval_start + 5}')
        prep_position = self.ws[f'K{approval_start + 5}']
        prep_position.value = "Manager"
        prep_position.font = Font(name='Arial', size=10)
        prep_position.alignment = Alignment(horizontal='center', vertical='center')
        
        # Department
        self.ws.merge_cells(f'K{approval_start + 6}:M{approval_start + 6}')
        prep_dept = self.ws[f'K{approval_start + 6}']
        prep_dept.value = "HR/ADMIN Department"
        prep_dept.font = Font(name='Arial', size=9, italic=True)
        prep_dept.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add final spacing
        self.ws.row_dimensions[approval_start + 7].height = 20
    
    def freeze_panes(self):
        """Freeze header rows and first columns"""
        self.ws.freeze_panes = 'D9'  # Freeze first 3 columns and 8 rows
    
    def add_print_settings(self):
        """Configure print settings"""
        from openpyxl.worksheet.page import PageMargins
        
        # Set margins
        self.ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
        
        # Set to landscape
        self.ws.page_setup.orientation = self.ws.ORIENTATION_LANDSCAPE
        
        # Fit to page
        self.ws.page_setup.fitToWidth = 1
        self.ws.page_setup.fitToHeight = 0
        
        # Set print area (adjust based on actual columns)
        # self.ws.print_area = f'A1:AN{self.ws.max_row}'
        
        # Repeat rows 1-8 on each page
        self.ws.print_title_rows = '1:8'
    
    def save(self):
        """Save the workbook with all formatting applied"""
        self.write_headers()
        self.write_data()
        self.set_column_widths()
        self.apply_department_colors()
        self.apply_borders()
        self.add_signatures()
        self.freeze_panes()
        self.add_print_settings()
        
        # Add summary sheets if database is available
        if self.dbase_df is not None:
            self.add_cost_center_summary()
            self.add_cash_cost_center_summary()  # NEW: Add cash-only summary
            self.add_cash_payroll_list()
        
        # Save workbook
        self.wb.save(self.filename)
        return self.filename
    
    def add_cost_center_summary(self):
        """Add Cost Center Summary sheet with detailed breakdown"""
        ws_summary = self.wb.create_sheet("Cost Center Summary")
        
        # Header
        ws_summary.merge_cells('A1:AI1')
        ws_summary['A1'] = COMPANY_NAME
        ws_summary['A1'].font = Font(name='Arial', size=14, bold=True, color='C00000')
        ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_summary.row_dimensions[1].height = 25
        
        ws_summary.merge_cells('A2:AI2')
        ws_summary['A2'] = "COST CENTER SUMMARY - DETAILED BREAKDOWN"
        ws_summary['A2'].font = Font(name='Arial', size=12, bold=True)
        ws_summary['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws_summary.row_dimensions[2].height = 22
        
        month_info = MONTH_CONFIG[self.month]
        if self.cutoff == 'first':
            period_text = f"{month_info['prev']} 26 - {self.month} 10"
            cutoff_text = f"{self.month} 15"
        else:
            period_text = f"{self.month} 11 - 25"
            cutoff_text = f"{self.month} {month_info['days']}"
        
        ws_summary.merge_cells('A3:AI3')
        ws_summary['A3'] = f"Period: {period_text}, {YEAR} | Cutoff: {cutoff_text}, {YEAR}"
        ws_summary['A3'].font = Font(name='Arial', size=10)
        ws_summary['A3'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws_summary.row_dimensions[4].height = 15
        
        # Column headers - All requested columns
        headers = [
            'CCR CODE',          # A
            'Cost Center',       # B
            'Emp Count',         # C
            'Basic Salary',      # D
            'OT A',              # E
            'OT B',              # F
            'OT C',              # G
            'SIL',               # H
            'Other Taxable Earnings',  # I
            'Total Lates/UT',    # J
            'Total Absences',    # K
            'Other Deduct (Sal Adj)',  # L
            'Total Deduct',      # M
            'SSS EE',            # N
            'PHEALTH EE',        # O
            'PAG-IBIG EE',       # P
            'TOT YEE Contri',    # Q
            'Statutory MWE',     # R
            'Taxable Compensation',  # S
            'NT Other Earnings (13th Month)',  # T
            'NT Other Earnings (Pos Allow)',  # U
            'NT Other Earnings (SIL Conv)',  # V
            'Other Non-Taxable Compensation',  # W
            'Total Compensation',  # X
            'Other Deduct (CoMat/Med Fee)',  # Y
            'SSS Loan',          # Z
            'Pag-ibig Loan',     # AA
            'HMI Membership',    # AB
            'Tax',               # AC
            'Net Pay',           # AD
            'SSS ER',            # AE
            'ECC',               # AF
            'PHEALTH ER',        # AG
            'Pag-ibig ER',       # AH
            '13TH_MONTH'         # AI
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_summary.cell(row=5, column=col_idx, value=header)
            cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        ws_summary.row_dimensions[5].height = 40
        
        # Collect cost center data from main sheet
        ccr_summary = {}
        
        for idx, row in self.df.iterrows():
            ccr_code = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
            row_name = str(row.iloc[2]) if len(row) > 2 else ''
            
            # Skip empty rows and grand total
            if not ccr_code or 'GRAND TOTAL' in row_name:
                continue
            
            # Map CCR code to name
            ccr_name = CCR_CODE_MAPPING.get(ccr_code, ccr_code)
            
            # Identify department totals
            if 'TOTAL' in row_name and ccr_code:
                if ccr_name not in ccr_summary:
                    ccr_summary[ccr_name] = {
                        'code': ccr_code,
                        'emp_count': safe_int(row.iloc[1]) if len(row) > 1 else 0,
                        # Basic salary and earnings
                        'basic': safe_float(row.iloc[7]) if len(row) > 7 else 0,  # H
                        'ot_a': safe_float(row.iloc[8]) if len(row) > 8 else 0,  # I
                        'ot_b': safe_float(row.iloc[9]) if len(row) > 9 else 0,  # J
                        'ot_c': safe_float(row.iloc[10]) if len(row) > 10 else 0,  # K
                        'sil': safe_float(row.iloc[11]) if len(row) > 11 else 0,  # L
                        'other_taxable': safe_float(row.iloc[12]) if len(row) > 12 else 0,  # M
                        # Deductions
                        'total_lates': safe_float(row.iloc[13]) if len(row) > 13 else 0,  # N
                        'total_absences': safe_float(row.iloc[14]) if len(row) > 14 else 0,  # O
                        'other_deduct_sal': safe_float(row.iloc[15]) if len(row) > 15 else 0,  # P
                        'total_deduct': safe_float(row.iloc[16]) if len(row) > 16 else 0,  # Q
                        # Employee contributions
                        'sss_ee': safe_float(row.iloc[17]) if len(row) > 17 else 0,  # R
                        'phealth_ee': safe_float(row.iloc[18]) if len(row) > 18 else 0,  # S
                        'pagibig_ee': safe_float(row.iloc[19]) if len(row) > 19 else 0,  # T
                        'tot_yee_contri': safe_float(row.iloc[20]) if len(row) > 20 else 0,  # U
                        'statutory_mwe': safe_float(row.iloc[21]) if len(row) > 21 else 0,  # V
                        'taxable_comp': safe_float(row.iloc[22]) if len(row) > 22 else 0,  # W
                        # Non-taxable earnings
                        'nt_13th': safe_float(row.iloc[23]) if len(row) > 23 else 0,  # X
                        'nt_pos_allow': safe_float(row.iloc[24]) if len(row) > 24 else 0,  # Y
                        'nt_sil_conv': safe_float(row.iloc[25]) if len(row) > 25 else 0,  # Z
                        'other_nt_comp': safe_float(row.iloc[26]) if len(row) > 26 else 0,  # AA
                        'total_comp': safe_float(row.iloc[27]) if len(row) > 27 else 0,  # AB
                        # Other deductions
                        'other_deduct_comat': safe_float(row.iloc[28]) if len(row) > 28 else 0,  # AC
                        'sss_loan': safe_float(row.iloc[29]) if len(row) > 29 else 0,  # AD
                        'pagibig_loan': safe_float(row.iloc[30]) if len(row) > 30 else 0,  # AE
                        'hmi_membership': safe_float(row.iloc[31]) if len(row) > 31 else 0,  # AF
                        'tax': safe_float(row.iloc[32]) if len(row) > 32 else 0,  # AG
                        'net_pay': safe_float(row.iloc[33]) if len(row) > 33 else 0,  # AH
                        # Employer contributions
                        'sss_er': safe_float(row.iloc[34]) if len(row) > 34 else 0,  # AI
                        'ecc': safe_float(row.iloc[35]) if len(row) > 35 else 0,  # AJ
                        'phealth_er': safe_float(row.iloc[36]) if len(row) > 36 else 0,  # AK
                        'pagibig_er': safe_float(row.iloc[37]) if len(row) > 37 else 0,  # AL
                        '13th_month': safe_float(row.iloc[38]) if len(row) > 38 else 0,  # AM
                    }
        
        # Write data in order
        row_idx = 6
        ccr_order = ['IND2001', 'IND2005', 'IND2101', 'IND2102', 'IND0202', 'IND0202-1', 
                    'IND0203', 'IND0203-1', 'IND0204', 'IND0205', 'IND0503', 'IND0506',
                    'IND0702', 'D2001', 'D2005', 'IND1002']
        
        for ccr_name in ccr_order:
            if ccr_name in ccr_summary:
                data = ccr_summary[ccr_name]
                
                # Write all data columns
                col_data = [
                    data['code'],           # A - CCR CODE
                    ccr_name,               # B - Cost Center
                    data['emp_count'],      # C - Emp Count
                    data['basic'],          # D - Basic Salary
                    data['ot_a'],           # E - OT A
                    data['ot_b'],           # F - OT B
                    data['ot_c'],           # G - OT C
                    data['sil'],            # H - SIL
                    data['other_taxable'],  # I - Other Taxable Earnings
                    data['total_lates'],    # J - Total Lates/UT
                    data['total_absences'], # K - Total Absences
                    data['other_deduct_sal'], # L - Other Deduct (Sal Adj)
                    data['total_deduct'],   # M - Total Deduct
                    data['sss_ee'],         # N - SSS EE
                    data['phealth_ee'],     # O - PHEALTH EE
                    data['pagibig_ee'],     # P - PAG-IBIG EE
                    data['tot_yee_contri'], # Q - TOT YEE Contri
                    data['statutory_mwe'],  # R - Statutory MWE
                    data['taxable_comp'],   # S - Taxable Compensation
                    data['nt_13th'],        # T - NT Other Earnings (13th Month)
                    data['nt_pos_allow'],   # U - NT Other Earnings (Pos Allow)
                    data['nt_sil_conv'],    # V - NT Other Earnings (SIL Conv)
                    data['other_nt_comp'],  # W - Other Non-Taxable Compensation
                    data['total_comp'],     # X - Total Compensation
                    data['other_deduct_comat'], # Y - Other Deduct (CoMat/Med Fee)
                    data['sss_loan'],       # Z - SSS Loan
                    data['pagibig_loan'],   # AA - Pag-ibig Loan
                    data['hmi_membership'], # AB - HMI Membership
                    data['tax'],            # AC - Tax
                    data['net_pay'],        # AD - Net Pay
                    data['sss_er'],         # AE - SSS ER
                    data['ecc'],            # AF - ECC
                    data['phealth_er'],     # AG - PHEALTH ER
                    data['pagibig_er'],     # AH - Pag-ibig ER
                    data['13th_month']      # AI - 13TH_MONTH
                ]
                
                for col_idx, value in enumerate(col_data, start=1):
                    cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Format based on column type
                    if col_idx == 1:  # CCR CODE
                        cell.font = Font(name='Arial', size=9, bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    elif col_idx == 2:  # Cost Center
                        cell.font = Font(name='Arial', size=9, bold=True)
                    elif col_idx == 3:  # Emp Count
                        cell.alignment = Alignment(horizontal='center')
                    elif col_idx >= 4:  # All numeric columns
                        if value != 0:
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right')
                    
                    # Apply borders and alternating colors
                    fill_color = 'F2F2F2' if row_idx % 2 == 0 else 'FFFFFF'
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                
                ws_summary.row_dimensions[row_idx].height = 20
                row_idx += 1
        
        # Grand total row
        total_start_row = 6
        total_end_row = row_idx - 1
        
        ws_summary.cell(row=row_idx, column=1, value='').font = Font(name='Arial', size=10, bold=True)
        ws_summary.cell(row=row_idx, column=2, value='GRAND TOTAL').font = Font(name='Arial', size=10, bold=True)
        ws_summary.cell(row=row_idx, column=3, value=f'=SUM(C{total_start_row}:C{total_end_row})')
        
        # Add formulas for all numeric columns
        numeric_columns = list(range(4, len(headers) + 1))  # Columns D to AI
        
        for col_idx in numeric_columns:
            if col_idx <= len(headers):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                formula = f'=SUM({col_letter}{total_start_row}:{col_letter}{total_end_row})'
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=formula)
                cell.font = Font(name='Arial', size=10, bold=True, color='C00000')
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                cell.border = Border(
                    left=Side(style='medium'), right=Side(style='medium'),
                    top=Side(style='double'), bottom=Side(style='double')
                )
        
        ws_summary.row_dimensions[row_idx].height = 25
        
        # Set column widths
        column_widths = {
            'A': 8,   # CCR CODE
            'B': 15,  # Cost Center
            'C': 10,  # Emp Count
            'D': 12,  # Basic Salary
            'E': 8,   # OT A
            'F': 8,   # OT B
            'G': 8,   # OT C
            'H': 8,   # SIL
            'I': 12,  # Other Taxable Earnings
            'J': 12,  # Total Lates/UT
            'K': 12,  # Total Absences
            'L': 15,  # Other Deduct (Sal Adj)
            'M': 12,  # Total Deduct
            'N': 10,  # SSS EE
            'O': 12,  # PHEALTH EE
            'P': 12,  # PAG-IBIG EE
            'Q': 12,  # TOT YEE Contri
            'R': 12,  # Statutory MWE
            'S': 15,  # Taxable Compensation
            'T': 15,  # NT Other Earnings (13th Month)
            'U': 15,  # NT Other Earnings (Pos Allow)
            'V': 15,  # NT Other Earnings (SIL Conv)
            'W': 18,  # Other Non-Taxable Compensation
            'X': 12,  # Total Compensation
            'Y': 18,  # Other Deduct (CoMat/Med Fee)
            'Z': 10,  # SSS Loan
            'AA': 12, # Pag-ibig Loan
            'AB': 12, # HMI Membership
            'AC': 10, # Tax
            'AD': 12, # Net Pay
            'AE': 10, # SSS ER
            'AF': 10, # ECC
            'AG': 12, # PHEALTH ER
            'AH': 12, # Pag-ibig ER
            'AI': 12  # 13TH_MONTH
        }
        
        for col_letter, width in column_widths.items():
            ws_summary.column_dimensions[col_letter].width = width
        
        # Freeze panes (headers and first columns)
        ws_summary.freeze_panes = 'D6'
        
        # Add print settings
        ws_summary.page_setup.orientation = ws_summary.ORIENTATION_LANDSCAPE
        ws_summary.page_setup.fitToWidth = 1
        ws_summary.page_setup.fitToHeight = 0
        ws_summary.print_title_rows = '1:5'  # Repeat headers
    
    def add_cash_cost_center_summary(self):
        """Add Cost Center Summary sheet specifically for CASH payroll employees only"""
        ws_cash_summary = self.wb.create_sheet("Cost Center Summary (CASH)")
        
        # Header
        ws_cash_summary.merge_cells('A1:AI1')
        ws_cash_summary['A1'] = COMPANY_NAME
        ws_cash_summary['A1'].font = Font(name='Arial', size=14, bold=True, color='C00000')
        ws_cash_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_cash_summary.row_dimensions[1].height = 25
        
        ws_cash_summary.merge_cells('A2:AI2')
        ws_cash_summary['A2'] = "COST CENTER SUMMARY - DETAILED BREAKDOWN (CASH PAYROLL ONLY)"
        ws_cash_summary['A2'].font = Font(name='Arial', size=12, bold=True)
        ws_cash_summary['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws_cash_summary.row_dimensions[2].height = 22
        
        month_info = MONTH_CONFIG[self.month]
        if self.cutoff == 'first':
            period_text = f"{month_info['prev']} 26 - {self.month} 10"
            cutoff_text = f"{self.month} 15"
        else:
            period_text = f"{self.month} 11 - 25"
            cutoff_text = f"{self.month} {month_info['days']}"
        
        ws_cash_summary.merge_cells('A3:AI3')
        ws_cash_summary['A3'] = f"Period: {period_text}, {YEAR} | Cutoff: {cutoff_text}, {YEAR}"
        ws_cash_summary['A3'].font = Font(name='Arial', size=10)
        ws_cash_summary['A3'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws_cash_summary.row_dimensions[4].height = 15
        
        # Column headers - Same as regular summary but for cash employees only
        headers = [
            'CCR CODE',          # A
            'Cost Center',       # B
            'Cash Emp Count',    # C - Changed to Cash Emp Count
            'Basic Salary',      # D
            'OT A',              # E
            'OT B',              # F
            'OT C',              # G
            'SIL',               # H
            'Other Taxable Earnings',  # I
            'Total Lates/UT',    # J
            'Total Absences',    # K
            'Other Deduct (Sal Adj)',  # L
            'Total Deduct',      # M
            'SSS EE',            # N
            'PHEALTH EE',        # O
            'PAG-IBIG EE',       # P
            'TOT YEE Contri',    # Q
            'Statutory MWE',     # R
            'Taxable Compensation',  # S
            'NT Other Earnings (13th Month)',  # T
            'NT Other Earnings (Pos Allow)',  # U
            'NT Other Earnings (SIL Conv)',  # V
            'Other Non-Taxable Compensation',  # W
            'Total Compensation',  # X
            'Other Deduct (CoMat/Med Fee)',  # Y
            'SSS Loan',          # Z
            'Pag-ibig Loan',     # AA
            'HMI Membership',    # AB
            'Tax',               # AC
            'Net Pay',           # AD
            'SSS ER',            # AE
            'ECC',               # AF
            'PHEALTH ER',        # AG
            'Pag-ibig ER',       # AH
            '13TH_MONTH'         # AI
        ]
        
        # Write headers with different color for cash summary
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_cash_summary.cell(row=5, column=col_idx, value=header)
            cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')  # Orange for cash
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        ws_cash_summary.row_dimensions[5].height = 40
        
        # Create account lookup to identify cash employees
        account_lookup = {}
        if self.dbase_df is not None:
            for idx, row in self.dbase_df.iterrows():
                emp_id = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
                if emp_id and emp_id.replace('.', '').replace('-', '').isdigit():
                    emp_id_clean = emp_id.split('.')[0]
                    # Check column 3 for account number
                    if len(row) > 3 and pd.notna(row.iloc[3]):
                        acct_val = row.iloc[3]
                        if isinstance(acct_val, (int, float)):
                            account_no = str(int(acct_val)) if pd.notna(acct_val) else None
                        else:
                            account_no = str(acct_val) if pd.notna(acct_val) else None
                        
                        # Clean account number - keep only digits
                        if account_no:
                            account_clean = ''.join(filter(str.isdigit, account_no))
                            if len(account_clean) >= 10:
                                account_lookup[emp_id_clean] = account_clean
                            else:
                                account_lookup[emp_id_clean] = None  # Mark as cash employee
                        else:
                            account_lookup[emp_id_clean] = None  # Mark as cash employee
                    else:
                        account_lookup[emp_id_clean] = None  # Mark as cash employee
        
        # Collect cash payroll data by cost center
        ccr_cash_summary = {}
        
        for idx, row in self.df.iterrows():
            ccr_code = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
            emp_id = str(row.iloc[1]) if len(row) > 1 and pd.notna(row.iloc[1]) else ''
            row_name = str(row.iloc[2]) if len(row) > 2 else ''
            
            # Skip total rows, grand total, and non-employee rows
            if not ccr_code or 'TOTAL' in row_name or 'GRAND' in row_name:
                continue
            
            # Skip if not a valid employee ID
            if not emp_id.replace('.', '').replace('-', '').isdigit():
                continue
            
            # Check if employee is cash payroll
            emp_id_clean = emp_id.split('.')[0]
            is_cash_employee = False
            
            if emp_id_clean in account_lookup:
                has_account = account_lookup[emp_id_clean]
                is_cash_employee = (has_account is None or has_account == '')
            else:
                # If not in database, assume cash employee
                is_cash_employee = True
            
            if not is_cash_employee:
                continue  # Skip bank employees
            
            # Map CCR code to name
            ccr_name = CCR_CODE_MAPPING.get(ccr_code, ccr_code)
            
            # Initialize cost center if not exists
            if ccr_name not in ccr_cash_summary:
                ccr_cash_summary[ccr_name] = {
                    'code': ccr_code,
                    'cash_emp_count': 0,
                    # Basic salary and earnings
                    'basic': 0,
                    'ot_a': 0,
                    'ot_b': 0,
                    'ot_c': 0,
                    'sil': 0,
                    'other_taxable': 0,
                    # Deductions
                    'total_lates': 0,
                    'total_absences': 0,
                    'other_deduct_sal': 0,
                    'total_deduct': 0,
                    # Employee contributions
                    'sss_ee': 0,
                    'phealth_ee': 0,
                    'pagibig_ee': 0,
                    'tot_yee_contri': 0,
                    'statutory_mwe': 0,
                    'taxable_comp': 0,
                    # Non-taxable earnings
                    'nt_13th': 0,
                    'nt_pos_allow': 0,
                    'nt_sil_conv': 0,
                    'other_nt_comp': 0,
                    'total_comp': 0,
                    # Other deductions
                    'other_deduct_comat': 0,
                    'sss_loan': 0,
                    'pagibig_loan': 0,
                    'hmi_membership': 0,
                    'tax': 0,
                    'net_pay': 0,
                    # Employer contributions
                    'sss_er': 0,
                    'ecc': 0,
                    'phealth_er': 0,
                    'pagibig_er': 0,
                    '13th_month': 0
                }
            
            # Add cash employee to count
            ccr_cash_summary[ccr_name]['cash_emp_count'] += 1
            
            # Accumulate values from employee row
            try:
                # Basic salary and earnings
                ccr_cash_summary[ccr_name]['basic'] += safe_float(row.iloc[7]) if len(row) > 7 else 0  # H
                ccr_cash_summary[ccr_name]['ot_a'] += safe_float(row.iloc[8]) if len(row) > 8 else 0  # I
                ccr_cash_summary[ccr_name]['ot_b'] += safe_float(row.iloc[9]) if len(row) > 9 else 0  # J
                ccr_cash_summary[ccr_name]['ot_c'] += safe_float(row.iloc[10]) if len(row) > 10 else 0  # K
                ccr_cash_summary[ccr_name]['sil'] += safe_float(row.iloc[11]) if len(row) > 11 else 0  # L
                ccr_cash_summary[ccr_name]['other_taxable'] += safe_float(row.iloc[12]) if len(row) > 12 else 0  # M
                # Deductions
                ccr_cash_summary[ccr_name]['total_lates'] += safe_float(row.iloc[13]) if len(row) > 13 else 0  # N
                ccr_cash_summary[ccr_name]['total_absences'] += safe_float(row.iloc[14]) if len(row) > 14 else 0  # O
                ccr_cash_summary[ccr_name]['other_deduct_sal'] += safe_float(row.iloc[15]) if len(row) > 15 else 0  # P
                ccr_cash_summary[ccr_name]['total_deduct'] += safe_float(row.iloc[16]) if len(row) > 16 else 0  # Q
                # Employee contributions
                ccr_cash_summary[ccr_name]['sss_ee'] += safe_float(row.iloc[17]) if len(row) > 17 else 0  # R
                ccr_cash_summary[ccr_name]['phealth_ee'] += safe_float(row.iloc[18]) if len(row) > 18 else 0  # S
                ccr_cash_summary[ccr_name]['pagibig_ee'] += safe_float(row.iloc[19]) if len(row) > 19 else 0  # T
                ccr_cash_summary[ccr_name]['tot_yee_contri'] += safe_float(row.iloc[20]) if len(row) > 20 else 0  # U
                ccr_cash_summary[ccr_name]['statutory_mwe'] += safe_float(row.iloc[21]) if len(row) > 21 else 0  # V
                ccr_cash_summary[ccr_name]['taxable_comp'] += safe_float(row.iloc[22]) if len(row) > 22 else 0  # W
                # Non-taxable earnings
                ccr_cash_summary[ccr_name]['nt_13th'] += safe_float(row.iloc[23]) if len(row) > 23 else 0  # X
                ccr_cash_summary[ccr_name]['nt_pos_allow'] += safe_float(row.iloc[24]) if len(row) > 24 else 0  # Y
                ccr_cash_summary[ccr_name]['nt_sil_conv'] += safe_float(row.iloc[25]) if len(row) > 25 else 0  # Z
                ccr_cash_summary[ccr_name]['other_nt_comp'] += safe_float(row.iloc[26]) if len(row) > 26 else 0  # AA
                ccr_cash_summary[ccr_name]['total_comp'] += safe_float(row.iloc[27]) if len(row) > 27 else 0  # AB
                # Other deductions
                ccr_cash_summary[ccr_name]['other_deduct_comat'] += safe_float(row.iloc[28]) if len(row) > 28 else 0  # AC
                ccr_cash_summary[ccr_name]['sss_loan'] += safe_float(row.iloc[29]) if len(row) > 29 else 0  # AD
                ccr_cash_summary[ccr_name]['pagibig_loan'] += safe_float(row.iloc[30]) if len(row) > 30 else 0  # AE
                ccr_cash_summary[ccr_name]['hmi_membership'] += safe_float(row.iloc[31]) if len(row) > 31 else 0  # AF
                ccr_cash_summary[ccr_name]['tax'] += safe_float(row.iloc[32]) if len(row) > 32 else 0  # AG
                ccr_cash_summary[ccr_name]['net_pay'] += safe_float(row.iloc[33]) if len(row) > 33 else 0  # AH
                # Employer contributions
                ccr_cash_summary[ccr_name]['sss_er'] += safe_float(row.iloc[34]) if len(row) > 34 else 0  # AI
                ccr_cash_summary[ccr_name]['ecc'] += safe_float(row.iloc[35]) if len(row) > 35 else 0  # AJ
                ccr_cash_summary[ccr_name]['phealth_er'] += safe_float(row.iloc[36]) if len(row) > 36 else 0  # AK
                ccr_cash_summary[ccr_name]['pagibig_er'] += safe_float(row.iloc[37]) if len(row) > 37 else 0  # AL
                ccr_cash_summary[ccr_name]['13th_month'] += safe_float(row.iloc[38]) if len(row) > 38 else 0  # AM
            except Exception as e:
                print(f"Error accumulating cash data for {ccr_name}: {e}")
                continue
        
        # Write data in order
        row_idx = 6
        ccr_order = ['IND2001', 'IND2005', 'IND2101', 'IND2102', 'IND0202', 'IND0202-1', 
                    'IND0203', 'IND0203-1', 'IND0204', 'IND0205', 'IND0503', 'IND0506',
                    'IND0702', 'D2001', 'D2005', 'IND1002']
        
        total_cash_employees = 0
        total_cash_net_pay = 0
        
        for ccr_name in ccr_order:
            if ccr_name in ccr_cash_summary:
                data = ccr_cash_summary[ccr_name]
                
                total_cash_employees += data['cash_emp_count']
                total_cash_net_pay += data['net_pay']
                
                # Write all data columns
                col_data = [
                    data['code'],           # A - CCR CODE
                    ccr_name,               # B - Cost Center
                    data['cash_emp_count'], # C - Cash Emp Count
                    data['basic'],          # D - Basic Salary
                    data['ot_a'],           # E - OT A
                    data['ot_b'],           # F - OT B
                    data['ot_c'],           # G - OT C
                    data['sil'],            # H - SIL
                    data['other_taxable'],  # I - Other Taxable Earnings
                    data['total_lates'],    # J - Total Lates/UT
                    data['total_absences'], # K - Total Absences
                    data['other_deduct_sal'], # L - Other Deduct (Sal Adj)
                    data['total_deduct'],   # M - Total Deduct
                    data['sss_ee'],         # N - SSS EE
                    data['phealth_ee'],     # O - PHEALTH EE
                    data['pagibig_ee'],     # P - PAG-IBIG EE
                    data['tot_yee_contri'], # Q - TOT YEE Contri
                    data['statutory_mwe'],  # R - Statutory MWE
                    data['taxable_comp'],   # S - Taxable Compensation
                    data['nt_13th'],        # T - NT Other Earnings (13th Month)
                    data['nt_pos_allow'],   # U - NT Other Earnings (Pos Allow)
                    data['nt_sil_conv'],    # V - NT Other Earnings (SIL Conv)
                    data['other_nt_comp'],  # W - Other Non-Taxable Compensation
                    data['total_comp'],     # X - Total Compensation
                    data['other_deduct_comat'], # Y - Other Deduct (CoMat/Med Fee)
                    data['sss_loan'],       # Z - SSS Loan
                    data['pagibig_loan'],   # AA - Pag-ibig Loan
                    data['hmi_membership'], # AB - HMI Membership
                    data['tax'],            # AC - Tax
                    data['net_pay'],        # AD - Net Pay
                    data['sss_er'],         # AE - SSS ER
                    data['ecc'],            # AF - ECC
                    data['phealth_er'],     # AG - PHEALTH ER
                    data['pagibig_er'],     # AH - Pag-ibig ER
                    data['13th_month']      # AI - 13TH_MONTH
                ]
                
                for col_idx, value in enumerate(col_data, start=1):
                    cell = ws_cash_summary.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Format based on column type
                    if col_idx == 1:  # CCR CODE
                        cell.font = Font(name='Arial', size=9, bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    elif col_idx == 2:  # Cost Center
                        cell.font = Font(name='Arial', size=9, bold=True)
                    elif col_idx == 3:  # Cash Emp Count
                        cell.alignment = Alignment(horizontal='center')
                    elif col_idx >= 4:  # All numeric columns
                        if value != 0:
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right')
                    
                    # Apply borders and alternating colors with orange tint
                    fill_color = 'FFE6CC' if row_idx % 2 == 0 else 'FFF2E6'  # Light orange shades
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                
                ws_cash_summary.row_dimensions[row_idx].height = 20
                row_idx += 1
        
        # Grand total row for cash payroll
        total_start_row = 6
        total_end_row = row_idx - 1
        
        if total_end_row >= total_start_row:  # Only add total if there's data
            ws_cash_summary.cell(row=row_idx, column=1, value='').font = Font(name='Arial', size=10, bold=True)
            ws_cash_summary.cell(row=row_idx, column=2, value='CASH PAYROLL TOTAL').font = Font(name='Arial', size=10, bold=True, color='C00000')
            ws_cash_summary.cell(row=row_idx, column=3, value=f'=SUM(C{total_start_row}:C{total_end_row})')
            
            # Add formulas for all numeric columns
            numeric_columns = list(range(4, len(headers) + 1))  # Columns D to AI
            
            for col_idx in numeric_columns:
                if col_idx <= len(headers):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    formula = f'=SUM({col_letter}{total_start_row}:{col_letter}{total_end_row})'
                    cell = ws_cash_summary.cell(row=row_idx, column=col_idx, value=formula)
                    cell.font = Font(name='Arial', size=10, bold=True, color='C00000')
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                    cell.fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')  # Darker orange
                    cell.border = Border(
                        left=Side(style='medium'), right=Side(style='medium'),
                        top=Side(style='double'), bottom=Side(style='double')
                    )
            
            # Add summary note
            summary_row = row_idx + 1
            ws_cash_summary.merge_cells(f'A{summary_row}:AI{summary_row}')
            ws_cash_summary.cell(row=summary_row, column=1, 
                               value=f'Summary: {total_cash_employees} cash employees | Total Cash Payroll: ₱{total_cash_net_pay:,.2f}')
            ws_cash_summary.cell(row=summary_row, column=1).font = Font(name='Arial', size=11, bold=True, color='FF9900')
            ws_cash_summary.cell(row=summary_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws_cash_summary.row_dimensions[summary_row].height = 25
        
        ws_cash_summary.row_dimensions[row_idx].height = 25
        
        # Set column widths
        column_widths = {
            'A': 8,   # CCR CODE
            'B': 15,  # Cost Center
            'C': 12,  # Cash Emp Count
            'D': 12,  # Basic Salary
            'E': 8,   # OT A
            'F': 8,   # OT B
            'G': 8,   # OT C
            'H': 8,   # SIL
            'I': 12,  # Other Taxable Earnings
            'J': 12,  # Total Lates/UT
            'K': 12,  # Total Absences
            'L': 15,  # Other Deduct (Sal Adj)
            'M': 12,  # Total Deduct
            'N': 10,  # SSS EE
            'O': 12,  # PHEALTH EE
            'P': 12,  # PAG-IBIG EE
            'Q': 12,  # TOT YEE Contri
            'R': 12,  # Statutory MWE
            'S': 15,  # Taxable Compensation
            'T': 15,  # NT Other Earnings (13th Month)
            'U': 15,  # NT Other Earnings (Pos Allow)
            'V': 15,  # NT Other Earnings (SIL Conv)
            'W': 18,  # Other Non-Taxable Compensation
            'X': 12,  # Total Compensation
            'Y': 18,  # Other Deduct (CoMat/Med Fee)
            'Z': 10,  # SSS Loan
            'AA': 12, # Pag-ibig Loan
            'AB': 12, # HMI Membership
            'AC': 10, # Tax
            'AD': 12, # Net Pay
            'AE': 10, # SSS ER
            'AF': 10, # ECC
            'AG': 12, # PHEALTH ER
            'AH': 12, # Pag-ibig ER
            'AI': 12  # 13TH_MONTH
        }
        
        for col_letter, width in column_widths.items():
            ws_cash_summary.column_dimensions[col_letter].width = width
        
        # Freeze panes (headers and first columns)
        ws_cash_summary.freeze_panes = 'D6'
        
        # Add print settings
        ws_cash_summary.page_setup.orientation = ws_cash_summary.ORIENTATION_LANDSCAPE
        ws_cash_summary.page_setup.fitToWidth = 1
        ws_cash_summary.page_setup.fitToHeight = 0
        ws_cash_summary.print_title_rows = '1:5'  # Repeat headers
        
        print(f"✓ Created Cash Cost Center Summary: {total_cash_employees} cash employees, Total: ₱{total_cash_net_pay:,.2f}")
    
    def add_cash_payroll_list(self):
        """Add Cash Payroll List sheet for employees without bank accounts"""
        ws_cash = self.wb.create_sheet("Cash Payroll")
        
        # Header
        ws_cash.merge_cells('A1:G1')
        ws_cash['A1'] = COMPANY_NAME
        ws_cash['A1'].font = Font(name='Arial', size=14, bold=True, color='C00000')
        ws_cash['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_cash.row_dimensions[1].height = 25
        
        ws_cash.merge_cells('A2:G2')
        ws_cash['A2'] = "CASH PAYROLL - Employees Without Bank Accounts"
        ws_cash['A2'].font = Font(name='Arial', size=12, bold=True)
        ws_cash['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws_cash.row_dimensions[2].height = 22
        
        month_info = MONTH_CONFIG[self.month]
        if self.cutoff == 'first':
            cutoff_text = f"{self.month} 15, {YEAR}"
        else:
            cutoff_text = f"{self.month} {month_info['days']}, {YEAR}"
        
        ws_cash.merge_cells('A3:G3')
        ws_cash['A3'] = f"Payroll Period: {cutoff_text}"
        ws_cash['A3'].font = Font(name='Arial', size=10)
        ws_cash['A3'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws_cash.row_dimensions[4].height = 15
        
        # Column headers
        headers = ['Cost Center', 'Emp ID', 'Employee Name', 'Position', 'Net Pay', 'Signature', 'Remarks']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_cash.cell(row=5, column=col_idx, value=header)
            cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        ws_cash.row_dimensions[5].height = 30
        
        # Create account lookup if dbase exists
        account_lookup = {}
        if self.dbase_df is not None:
            for idx, row in self.dbase_df.iterrows():
                emp_id = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
                if emp_id and emp_id.isdigit():
                    acct = row.iloc[3] if len(row) > 3 and pd.notna(row.iloc[3]) else None
                    account_lookup[emp_id] = acct
        
        # Get employees without bank accounts
        row_idx = 6
        cash_employees = []
        
        for idx, row in self.df.iterrows():
            ccr_code = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
            emp_id = str(row.iloc[1]) if len(row) > 1 and pd.notna(row.iloc[1]) else ''
            emp_name_parts = []
            
            # Skip total rows
            if 'TOTAL' in str(row.iloc[2]) or not emp_id.isdigit():
                continue
            
            # Check if employee has account
            has_account = account_lookup.get(emp_id)
            
            if not has_account or pd.isna(has_account):
                # Get employee details
                if len(row) > 2:
                    emp_name_parts.append(str(row.iloc[3]) if pd.notna(row.iloc[3]) else '')  # Last
                    emp_name_parts.append(str(row.iloc[4]) if len(row) > 4 and pd.notna(row.iloc[4]) else '')  # First
                    emp_name_parts.append(str(row.iloc[5]) if len(row) > 5 and pd.notna(row.iloc[5]) else '')  # Middle
                
                emp_name = ', '.join([p for p in emp_name_parts if p and p != 'nan'])
                position = "Daily Paid"
                net_pay = safe_float(row.iloc[33]) if len(row) > 33 else 0
                
                if net_pay > 0:
                    cash_employees.append({
                        'ccr': ccr_code,
                        'emp_id': emp_id,
                        'name': emp_name,
                        'position': position,
                        'net_pay': net_pay
                    })
        
        # Write cash employees
        for emp in sorted(cash_employees, key=lambda x: (x['ccr'], x['name'])):
            ws_cash.cell(row=row_idx, column=1, value=emp['ccr'])
            ws_cash.cell(row=row_idx, column=2, value=emp['emp_id']).alignment = Alignment(horizontal='center')
            ws_cash.cell(row=row_idx, column=3, value=emp['name'])
            ws_cash.cell(row=row_idx, column=4, value=emp['position'])
            ws_cash.cell(row=row_idx, column=5, value=emp['net_pay']).number_format = '₱#,##0.00'
            ws_cash.cell(row=row_idx, column=6, value='')  # Signature
            ws_cash.cell(row=row_idx, column=7, value='NO BANK ACCOUNT')
            
            # Apply formatting
            fill_color = 'FFF2CC' if row_idx % 2 == 0 else 'FFFFFF'
            for col in range(1, 8):
                cell = ws_cash.cell(row=row_idx, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                cell.font = Font(name='Arial', size=10)
            
            ws_cash.row_dimensions[row_idx].height = 22
            row_idx += 1
        
        # Total row
        if len(cash_employees) > 0:
            ws_cash.merge_cells(f'A{row_idx}:D{row_idx}')
            ws_cash.cell(row=row_idx, column=1, value=f'TOTAL CASH PAYROLL ({len(cash_employees)} Employees)')
            ws_cash.cell(row=row_idx, column=1).font = Font(name='Arial', size=11, bold=True)
            ws_cash.cell(row=row_idx, column=1).alignment = Alignment(horizontal='right', vertical='center')
            
            formula = f'=SUM(E6:E{row_idx-1})'
            ws_cash.cell(row=row_idx, column=5, value=formula)
            ws_cash.cell(row=row_idx, column=5).font = Font(name='Arial', size=11, bold=True, color='C00000')
            ws_cash.cell(row=row_idx, column=5).number_format = '₱#,##0.00'
            
            for col in range(1, 8):
                cell = ws_cash.cell(row=row_idx, column=col)
                cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                cell.border = Border(
                    left=Side(style='medium'), right=Side(style='medium'),
                    top=Side(style='double'), bottom=Side(style='double')
                )
        else:
            ws_cash.cell(row=row_idx, column=1, value='No employees without bank accounts')
            ws_cash.cell(row=row_idx, column=1).font = Font(name='Arial', size=10, italic=True)
        
        # Set column widths
        ws_cash.column_dimensions['A'].width = 12
        ws_cash.column_dimensions['B'].width = 10
        ws_cash.column_dimensions['C'].width = 30
        ws_cash.column_dimensions['D'].width = 15
        ws_cash.column_dimensions['E'].width = 15
        ws_cash.column_dimensions['F'].width = 25
        ws_cash.column_dimensions['G'].width = 20

# ============================================================================
# BDO SHEET CREATION HELPERS
# ============================================================================

def create_bdo_sheet(wb, sheet_name, df, total_employees, total_amount):
    """Create BDO bank payroll sheet"""
    ws = wb.create_sheet(sheet_name)
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "BANCO DE ORO (BDO) UNIBANK, INC."
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color='0033A0')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    ws.merge_cells('A2:E2')
    ws['A2'] = "Bank Payroll Crediting File"
    ws['A2'].font = Font(name='Arial', size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22
    
    ws.merge_cells('A3:E3')
    ws['A3'] = COMPANY_NAME
    ws['A3'].font = Font(name='Arial', size=11, bold=True)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A4:E4')
    ws['A4'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A4'].font = Font(name='Arial', size=9, italic=True)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[5].height = 10
    
    # Summary
    ws['A6'] = "Total Employees:"
    ws['A6'].font = Font(name='Arial', size=10, bold=True)
    ws['B6'] = total_employees
    ws['B6'].font = Font(name='Arial', size=10)
    
    ws['D6'] = "Total Amount:"
    ws['D6'].font = Font(name='Arial', size=10, bold=True)
    ws['D6'].alignment = Alignment(horizontal='right')
    ws['E6'] = total_amount
    ws['E6'].font = Font(name='Arial', size=10, bold=True, color='006100')
    ws['E6'].number_format = '₱#,##0.00'
    ws['E6'].alignment = Alignment(horizontal='right')
    
    ws.row_dimensions[7].height = 8
    
    # Headers
    headers = ['Account Number', 'Net Pay Amount', 'Employee Name', 'Status', 'Remarks']
    header_colors = ['4472C4', '4472C4', '4472C4', '70AD47', 'ED7D31']
    
    for col_idx, (header, color) in enumerate(zip(headers, header_colors), start=1):
        cell = ws.cell(row=8, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
    
    ws.row_dimensions[8].height = 35
    
    # Data
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=9):
        fill_color = 'F2F2F2' if row_idx % 2 == 0 else 'FFFFFF'
        
        ws.cell(row=row_idx, column=1, value=row_data[0]).font = Font(name='Courier New', size=10, bold=True)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=1).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        ws.cell(row=row_idx, column=2, value=row_data[1]).font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=row_idx, column=2).number_format = '₱#,##0.00'
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=row_idx, column=2).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        ws.cell(row=row_idx, column=3, value=row_data[2]).font = Font(name='Arial', size=10)
        ws.cell(row=row_idx, column=3).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        ws.cell(row=row_idx, column=4, value="Ready").font = Font(name='Arial', size=9, bold=True, color='006100')
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=4).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        ws.cell(row=row_idx, column=5, value="").fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
        
        ws.row_dimensions[row_idx].height = 22
    
    # Total row
    footer_row = len(df) + 9
    ws.row_dimensions[footer_row].height = 15
    footer_row += 1
    
    ws['A' + str(footer_row)] = "TOTAL:"
    ws['A' + str(footer_row)].font = Font(name='Arial', size=11, bold=True)
    ws['A' + str(footer_row)].alignment = Alignment(horizontal='right')
    ws['A' + str(footer_row)].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    ws.cell(row=footer_row, column=2, value=f'=SUM(B9:B{footer_row-2})')
    ws.cell(row=footer_row, column=2).font = Font(name='Arial', size=11, bold=True, color='C00000')
    ws.cell(row=footer_row, column=2).number_format = '₱#,##0.00'
    ws.cell(row=footer_row, column=2).alignment = Alignment(horizontal='right')
    ws.cell(row=footer_row, column=2).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    ws.merge_cells(f'C{footer_row}:E{footer_row}')
    ws.cell(row=footer_row, column=3, value=f"{total_employees} Employees")
    ws.cell(row=footer_row, column=3).font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=footer_row, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=footer_row, column=3).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    for col in range(1, 6):
        ws.cell(row=footer_row, column=col).border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='double'), bottom=Side(style='double')
        )
    
    ws.row_dimensions[footer_row].height = 28
    
    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    
    ws.freeze_panes = 'A9'

def create_cash_sheet(wb, sheet_name, df, total_employees, total_amount):
    """Create cash payroll sheet"""
    ws = wb.create_sheet(sheet_name)
    
    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = "CASH PAYROLL"
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color='C00000')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "Employees Without Bank Accounts"
    ws['A2'].font = Font(name='Arial', size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22
    
    ws.merge_cells('A3:F3')
    ws['A3'] = COMPANY_NAME
    ws['A3'].font = Font(name='Arial', size=11, bold=True)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A4:F4')
    ws['A4'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A4'].font = Font(name='Arial', size=9, italic=True)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[5].height = 10
    
    # Summary
    ws['A6'] = "Total Employees:"
    ws['A6'].font = Font(name='Arial', size=10, bold=True)
    ws['B6'] = total_employees
    ws['B6'].font = Font(name='Arial', size=10)
    
    ws['E6'] = "Total Cash:"
    ws['E6'].font = Font(name='Arial', size=10, bold=True)
    ws['E6'].alignment = Alignment(horizontal='right')
    ws['F6'] = total_amount
    ws['F6'].font = Font(name='Arial', size=10, bold=True, color='C00000')
    ws['F6'].number_format = '₱#,##0.00'
    ws['F6'].alignment = Alignment(horizontal='right')
    
    ws.row_dimensions[7].height = 8
    
    # Headers
    headers = ['Emp ID', 'Employee Name', 'Net Pay', 'Signature', 'Date Received', 'Remarks']
    header_color = 'C00000'
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
    
    ws.row_dimensions[8].height = 35
    
    # Data
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=9):
        fill_color = 'FFF2CC' if row_idx % 2 == 0 else 'FFFFFF'
        
        ws.cell(row=row_idx, column=1, value=row_data[0]).font = Font(name='Arial', size=10)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=1).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        ws.cell(row=row_idx, column=2, value=row_data[2]).font = Font(name='Arial', size=10)
        ws.cell(row=row_idx, column=2).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        ws.cell(row=row_idx, column=3, value=row_data[1]).font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=row_idx, column=3).number_format = '₱#,##0.00'
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal='right')
        ws.cell(row=row_idx, column=3).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        for col in [4, 5, 6]:
            ws.cell(row=row_idx, column=col, value="")
            ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        ws.cell(row=row_idx, column=6, value="NO BANK ACCOUNT")
        ws.cell(row=row_idx, column=6).font = Font(name='Arial', size=9, italic=True, color='C00000')
        
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
        
        ws.row_dimensions[row_idx].height = 25
    
    # Total row
    footer_row = len(df) + 9
    ws.row_dimensions[footer_row].height = 15
    footer_row += 1
    
    ws.merge_cells(f'A{footer_row}:B{footer_row}')
    ws.cell(row=footer_row, column=1, value="TOTAL CASH PAYROLL:")
    ws.cell(row=footer_row, column=1).font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=footer_row, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=footer_row, column=1).fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
    
    ws.cell(row=footer_row, column=3, value=f'=SUM(C9:C{footer_row-2})')
    ws.cell(row=footer_row, column=3).font = Font(name='Arial', size=11, bold=True, color='C00000')
    ws.cell(row=footer_row, column=3).number_format = '₱#,##0.00'
    ws.cell(row=footer_row, column=3).alignment = Alignment(horizontal='right')
    ws.cell(row=footer_row, column=3).fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
    
    ws.merge_cells(f'D{footer_row}:F{footer_row}')
    ws.cell(row=footer_row, column=4, value=f"{total_employees} Employees")
    ws.cell(row=footer_row, column=4).font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=footer_row, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=footer_row, column=4).fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
    
    for col in range(1, 7):
        ws.cell(row=footer_row, column=col).border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='double'), bottom=Side(style='double')
        )
    
    ws.row_dimensions[footer_row].height = 28
    
    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    
    ws.freeze_panes = 'A9'

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html', months=list(MONTH_CONFIG.keys()))

@app.route('/upload', methods=['POST'])
def upload_files():
    """Handle file uploads"""
    try:
        print("\n" + "="*80)
        print("PAYROLL PROCESSING STARTED")
        print("="*80)
        
        if 'payroll_file' not in request.files or 'dbase_file' not in request.files:
            print("ERROR: Missing required files in request")
            return jsonify({'error': 'Missing required files'}), 400
        
        payroll_file = request.files['payroll_file']
        dbase_file = request.files['dbase_file']
        month = request.form.get('month')
        cutoff = request.form.get('cutoff')
        
        print(f"Files received:")
        print(f"  - Payroll: {payroll_file.filename}")
        print(f"  - Database: {dbase_file.filename}")
        print(f"  - Month: {month}")
        print(f"  - Cutoff: {cutoff}")
        
        if not month or not cutoff:
            print("ERROR: Month or cutoff not provided")
            return jsonify({'error': 'Month and cutoff are required'}), 400
        
        if payroll_file.filename == '' or dbase_file.filename == '':
            print("ERROR: Empty filename(s)")
            return jsonify({'error': 'No files selected'}), 400
        
        if not (allowed_file(payroll_file.filename) and allowed_file(dbase_file.filename)):
            print("ERROR: Invalid file type(s)")
            return jsonify({'error': 'Invalid file type'}), 400
        
        # Save uploaded files
        payroll_filename = secure_filename(payroll_file.filename)
        dbase_filename = secure_filename(dbase_file.filename)
        
        payroll_path = os.path.join(app.config['UPLOAD_FOLDER'], payroll_filename)
        dbase_path = os.path.join(app.config['UPLOAD_FOLDER'], dbase_filename)
        
        print(f"\nSaving files:")
        print(f"  - Payroll: {payroll_path}")
        print(f"  - Database: {dbase_path}")
        
        payroll_file.save(payroll_path)
        dbase_file.save(dbase_path)
        print("Files saved successfully")
        
        # Read files
        print("\nReading Excel files...")
        try:
            payroll_df = pd.read_excel(payroll_path)
            print(f"✓ Payroll file read: {payroll_df.shape[0]} rows, {payroll_df.shape[1]} columns")
            print(f"  First 3 column names: {list(payroll_df.columns[:3])}")
        except Exception as e:
            print(f"ERROR reading payroll file: {e}")
            return jsonify({'error': f'Error reading payroll file: {str(e)}'}), 400
        
        try:
            dbase_df = pd.read_excel(dbase_path)
            print(f"✓ Database file read: {dbase_df.shape[0]} rows, {dbase_df.shape[1]} columns")
            print(f"  First 3 column names: {list(dbase_df.columns[:3])}")
        except Exception as e:
            print(f"ERROR reading database file: {e}")
            return jsonify({'error': f'Error reading database file: {str(e)}'}), 400
        
        # Process payroll
        print("\n" + "-"*80)
        print("STARTING PAYROLL PROCESSING")
        print("-"*80)
        
        try:
            processor = PayrollProcessor(payroll_df, dbase_df, month, cutoff)
            result_df = processor.process()
            print(f"\n✓ Payroll processing completed")
            print(f"  Result: {result_df.shape[0]} rows, {result_df.shape[1]} columns")
        except Exception as e:
            print(f"\nERROR in payroll processing:")
            print(f"  Error type: {type(e).__name__}")
            print(f"  Error message: {str(e)}")
            print(traceback.format_exc())
            return jsonify({'error': f'Processing error: {str(e)}'}), 500
        
        # Generate output filename
        month_code = MONTH_CONFIG[month]['code']
        cutoff_code = '15' if cutoff == 'first' else '30'
        output_filename = f"Payroll_{month_code}{cutoff_code}_{YEAR}.xlsx"
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        
        print(f"\nWriting output file: {output_filename}")
        
        # Write formatted Excel with summary sheets
        try:
            writer = FormattedExcelWriter(output_path, result_df, month, cutoff, dbase_df)
            writer.save()
            print(f"✓ Excel file written successfully")
        except Exception as e:
            print(f"\nERROR writing Excel file:")
            print(f"  Error type: {type(e).__name__}")
            print(f"  Error message: {str(e)}")
            print(traceback.format_exc())
            return jsonify({'error': f'Excel writing error: {str(e)}'}), 500
        
        print("\n" + "="*80)
        print("PAYROLL PROCESSING COMPLETED SUCCESSFULLY")
        print("="*80 + "\n")
        
        return jsonify({
            'success': True,
            'message': 'Payroll processed successfully',
            'download_url': f'/download/{output_filename}'
        })
        
    except Exception as e:
        print(f"\n{'='*80}")
        print("FATAL ERROR IN UPLOAD_FILES")
        print(f"{'='*80}")
        print(f"Error type: {type(e).__name__}")
        print(f"Error message: {str(e)}")
        print(f"\nFull traceback:")
        print(traceback.format_exc())
        print(f"{'='*80}\n")
        return jsonify({'error': str(e)}), 500

@app.route('/convert_bdo', methods=['POST'])
def convert_bdo():
    """Convert to BDO format"""
    try:
        if 'payroll_file' not in request.files or 'dbase_file' not in request.files:
            return jsonify({'error': 'Missing required files'}), 400
        
        payroll_file = request.files['payroll_file']
        dbase_file = request.files['dbase_file']
        
        # Save and process files
        payroll_filename = secure_filename(payroll_file.filename)
        dbase_filename = secure_filename(dbase_file.filename)
        
        payroll_path = os.path.join(app.config['UPLOAD_FOLDER'], payroll_filename)
        dbase_path = os.path.join(app.config['UPLOAD_FOLDER'], dbase_filename)
        
        payroll_file.save(payroll_path)
        dbase_file.save(dbase_path)
        
        # Read files - skip header rows for payroll file
        try:
            # Try to find the data start row by looking for "CCR" or numeric employee IDs
            payroll_df_temp = pd.read_excel(payroll_path, sheet_name=0, header=None)
            
            # Find the header row (look for "CCR CODE" or "ACCT NO" or numeric pattern)
            data_start_row = None
            for idx in range(min(20, len(payroll_df_temp))):
                row_vals = payroll_df_temp.iloc[idx].astype(str).str.upper()
                if any('CCR' in str(val) or 'ACCT' in str(val) or 'EMPLOYEE' in str(val) for val in row_vals):
                    data_start_row = idx + 1  # Data starts after header
                    break
            
            if data_start_row is None:
                # If no header found, look for first row with numeric employee ID
                for idx in range(min(20, len(payroll_df_temp))):
                    first_val = str(payroll_df_temp.iloc[idx, 0])
                    if first_val.isdigit() and len(first_val) == 6:  # Employee ID pattern
                        data_start_row = idx
                        break
            
            if data_start_row is None:
                data_start_row = 0
            
            print(f"Found data starting at row {data_start_row}")
            
            # Read again with correct starting row
            payroll_df = pd.read_excel(payroll_path, sheet_name=0, skiprows=data_start_row, header=None)
            
            # Check if first row still looks like headers, skip one more row
            first_row_str = ' '.join(str(v).upper() for v in payroll_df.iloc[0].tolist()[:5])
            if any(keyword in first_row_str for keyword in ['CCR', 'EMP', 'ACCT', 'NAME', 'SALARY', 'BASIC']):
                print(f"First row still has headers, skipping one more row")
                payroll_df = payroll_df.iloc[1:].reset_index(drop=True)
            
            print(f"Payroll file loaded: {len(payroll_df)} rows, {len(payroll_df.columns)} columns")
            print(f"First data row: {payroll_df.iloc[0].tolist()[:5]}")
            
        except Exception as e:
            return jsonify({'error': f'Error reading payroll file: {str(e)}'}), 400
        
        try:
            # Database file might also have headers
            dbase_df_temp = pd.read_excel(dbase_path, sheet_name=0, header=None)
            
            # Check if first row looks like headers
            first_row = dbase_df_temp.iloc[0].astype(str)
            if any(not str(val).isdigit() for val in first_row if pd.notna(val)):
                # Has headers, skip first row
                dbase_df = pd.read_excel(dbase_path, sheet_name=0, skiprows=1, header=None)
            else:
                dbase_df = dbase_df_temp
            
            print(f"Database file loaded: {len(dbase_df)} rows, {len(dbase_df.columns)} columns")
            print(f"First DB row: {dbase_df.iloc[0].tolist()[:5]}")
        except Exception as e:
            return jsonify({'error': f'Error reading database file: {str(e)}'}), 400
        
        # Validate minimum columns
        if len(payroll_df.columns) < 10:
            return jsonify({
                'error': f'Payroll file has insufficient columns. Expected at least 10 columns, found {len(payroll_df.columns)}'
            }), 400
        
        if len(dbase_df.columns) < 4:
            return jsonify({
                'error': f'Database file has insufficient columns. Expected at least 4 columns, found {len(dbase_df.columns)}'
            }), 400
        
        # Convert to BDO format
        try:
            converter = BDOConverter(payroll_df, dbase_df)
            result = converter.convert()
            bank_df = result['bank']
            cash_df = result['cash']
        except ValueError as ve:
            return jsonify({'error': str(ve)}), 400
        except Exception as e:
            return jsonify({'error': f'Conversion error: {str(e)}'}), 500
        
        if len(bank_df) == 0 and len(cash_df) == 0:
            return jsonify({'error': 'No valid employee records found after conversion'}), 400
        
        # Save result with enhanced formatting
        output_filename = f"BDO_Payroll_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        
        # Create workbook with multiple sheets
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # ===== SHEET 1: SUMMARY =====
        ws_summary = wb.create_sheet("Summary", 0)
        
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = "BANCO DE ORO (BDO) PAYROLL SUMMARY"
        ws_summary['A1'].font = Font(name='Arial', size=14, bold=True, color='0033A0')
        ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_summary.row_dimensions[1].height = 30
        
        ws_summary.merge_cells('A2:D2')
        ws_summary['A2'] = COMPANY_NAME
        ws_summary['A2'].font = Font(name='Arial', size=11, bold=True)
        ws_summary['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws_summary.merge_cells('A3:D3')
        ws_summary['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws_summary['A3'].font = Font(name='Arial', size=9, italic=True)
        ws_summary['A3'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws_summary.row_dimensions[5].height = 15
        
        # Summary table
        summary_data = [
            ['Payment Type', 'Employee Count', 'Total Amount', 'Percentage'],
            ['BDO Bank Payroll', result['bank_count'], result['bank_total'], (result['bank_total']/result['total']*100) if result['total'] > 0 else 0],
            ['Cash Payroll', result['cash_count'], result['cash_total'], (result['cash_total']/result['total']*100) if result['total'] > 0 else 0],
            ['TOTAL PAYROLL', result['bank_count'] + result['cash_count'], result['total'], 100]
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=6):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                
                if row_idx == 6:  # Header
                    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='0033A0', end_color='0033A0', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif row_idx == 9:  # Total
                    cell.font = Font(name='Arial', size=11, bold=True, color='C00000')
                    cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                else:
                    cell.font = Font(name='Arial', size=10)
                
                # Number formatting
                if col_idx == 3 and row_idx > 6:
                    cell.number_format = '₱#,##0.00'
                elif col_idx == 4 and row_idx > 6:
                    cell.number_format = '0.00%'
                    if isinstance(value, (int, float)):
                        cell.value = value / 100
                
                # Borders
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # Alignment
                if col_idx in [2, 3, 4]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 18
        ws_summary.column_dimensions['C'].width = 18
        ws_summary.column_dimensions['D'].width = 15
        
        # ===== SHEET 2: BDO BANK PAYROLL =====
        if len(bank_df) > 0:
            create_bdo_sheet(wb, "BDO Bank Payroll", bank_df, result['bank_count'], result['bank_total'])
        
        # ===== SHEET 3: CASH PAYROLL =====
        if len(cash_df) > 0:
            create_cash_sheet(wb, "Cash Payroll", cash_df, result['cash_count'], result['cash_total'])
        
        wb.save(output_path)
        
        return jsonify({
            'success': True,
            'message': f'BDO conversion completed - {result["bank_count"]} bank, {result["cash_count"]} cash',
            'download_url': f'/download/{output_filename}',
            'summary': {
                'bank_employees': result['bank_count'],
                'cash_employees': result['cash_count'],
                'bank_amount': f'₱{result["bank_total"]:,.2f}',
                'cash_amount': f'₱{result["cash_total"]:,.2f}',
                'total_amount': f'₱{result["total"]:,.2f}'
            }
        })
        
    except Exception as e:
        print(f"Error in convert_bdo: {traceback.format_exc()}")
        return jsonify({'error': f'Unexpected error: {str(e)}'}), 500

@app.route('/download/<filename>')
def download_file(filename):
    """Download processed file"""
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        return send_file(file_path, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.errorhandler(413)
def too_large(e):
    return jsonify({'error': 'File is too large. Maximum size is 16MB'}), 413

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=3553)